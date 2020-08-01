from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime


def funcao_criar_planilhas():
    # Função que cria as planilhas utilizadas na pasta
    workbook = Workbook('Planilha_Originaria_nova.xlsx')
    planilha1 = workbook.create_sheet('Operacoes')
    planilha2 = workbook.create_sheet('Tipo_do_Agente')
    workbook.save('Planilha_Originaria_nova.xlsx')
    return 'Nova planilha gerada'


# Saída
print(funcao_criar_planilhas())


def funcao_copiar_nova_plan1():
    # Função que copia as informações para a nova planilha (Operações)
    workbook1 = load_workbook('Planilha_Originaria.xlsx')
    work_planilha1 = workbook1['Operacoes']

    workbook2 = load_workbook('Planilha_Originaria_nova.xlsx')
    work_planilha2 = workbook2['Operacoes']

    max_linha = work_planilha1.max_row
    max_coluna = work_planilha1.max_column

    for i in range(1, max_linha + 1):
        for j in range(1, max_coluna + 1):
            cel = work_planilha1.cell(row=i, column=j)
            work_planilha2.cell(row=i, column=j + 4).value = cel.value

    return workbook2.save('Planilha_Originaria_nova.xlsx')


# Saída
funcao_copiar_nova_plan1()


def funcao_copiar_nova_plan2():
    # Função que copia as informações para a nova planilha (Tipo de Agente)
    workbook1 = load_workbook('Planilha_Originaria.xlsx')
    work_planilha1 = workbook1['Tipo_do_Agente']

    workbook2 = load_workbook('Planilha_Originaria_nova.xlsx')
    work_planilha2 = workbook2['Tipo_do_Agente']

    max_linha = work_planilha1.max_row
    max_coluna = work_planilha1.max_column

    for i in range(1, max_linha + 1):
        for j in range(1, max_coluna + 1):
            cel = work_planilha1.cell(row=i, column=j)
            work_planilha2.cell(row=i, column=j).value = cel.value

    return workbook2.save('Planilha_Originaria_nova.xlsx')


# Saída
funcao_copiar_nova_plan2()


def funcao_coluna_A():
    # Função que insere o nome do escritório responsável
    workbook2 = load_workbook('Planilha_Originaria_nova.xlsx')
    work_planilha2 = workbook2['Operacoes']

    work_planilha2['A1'] = 'Escritório Responsável'

    max_linha = work_planilha2.max_row

    for i in range(2, max_linha + 1):
        cel = "Regis Pontes Sociedade de Advogados / 220814"
        work_planilha2.cell(row=i, column=1).value = cel

    return workbook2.save('Planilha_Originaria_nova.xlsx')


# Saída
funcao_coluna_A()


def funcao_coluna_B():
    # Função que insere os números das pastas
    workbook2 = load_workbook('Planilha_Originaria_nova.xlsx')
    work_planilha2 = workbook2['Operacoes']

    work_planilha2['B1'] = 'Pasta'

    max_linha = work_planilha2.max_row

    for i in range(2, max_linha + 1):
        cel = work_planilha2.cell(row=i, column=7)
        work_planilha2.cell(row=i, column=2).value = cel.value
    work_planilha2.delete_cols(7)

    return workbook2.save('Planilha_Originaria_nova.xlsx')


# Saída
funcao_coluna_B()


def funcao_coluna_C():
    # Função que insere a situação
    workbook2 = load_workbook('Planilha_Originaria_nova.xlsx')
    work_planilha2 = workbook2['Operacoes']

    work_planilha2['C1'] = 'Situação'

    max_linha = work_planilha2.max_row

    for i in range(2, max_linha + 1):
        cel = "Pend.DCkJI-Resp.RP-SAR"
        work_planilha2.cell(row=i, column=3).value = cel

    return workbook2.save('Planilha_Originaria_nova.xlsx')


# Saída
funcao_coluna_C()


def funcao_coluna_D():
    # Função que insere a descrição do andamento
    workbook2 = load_workbook('Planilha_Originaria_nova.xlsx')
    work_planilha2 = workbook2['Operacoes']

    work_planilha2['D1'] = 'Descrição Andamento'

    max_linha = work_planilha2.max_row

    for i in range(2, max_linha + 1):
        cel = "Pendente de double check jurídico interno"
        work_planilha2.cell(row=i, column=4).value = cel

    return workbook2.save('Planilha_Originaria_nova.xlsx')


# Saída
funcao_coluna_D()


def funcao_coluna_N():
    # Função que insere o argumento (MWm)
    workbook2 = load_workbook('Planilha_Originaria_nova.xlsx')
    work_planilha2 = workbook2['Operacoes']

    max_linha = work_planilha2.max_row

    for i in range(2, max_linha + 1):
        cell_numero = float(work_planilha2.cell(row=i, column=14).value)
        cell_float_6d = f'{cell_numero:.6f}'
        cell = [str(cell_float_6d), 'MWn']
        work_planilha2.cell(row=i, column=14).value = ' '.join(cell)

    return workbook2.save('Planilha_Originaria_nova.xlsx')


# Saída
funcao_coluna_N()


def funcao_operacao():
    # Função que insere o termo (Termo de Cessão de Montantes) se o cliente for consumidor.
    workbook2 = load_workbook('Planilha_Originaria_nova.xlsx')
    work_planilha1 = workbook2['Operacoes']
    work_planilha2 = workbook2['Tipo_do_Agente']

    max_linha = work_planilha2.max_row

    for i in range(1, max_linha + 1):
        c1 = work_planilha2.cell(column=3, row=i)
        cel = "Termo de Cessão de Montantes"
        if c1.value == 'Consumidor':
            c2 = work_planilha1.cell(column=11, row=i)
            work_planilha1.cell(column=11, row=i).value = cel
    return workbook2.save('Planilha_Originaria_nova.xlsx')


# Saída
funcao_operacao()


def funcao_coluna_E():
    # Função que formata a data em (MM/YYYY)
    workbook2 = load_workbook('Planilha_Originaria_nova.xlsx')
    work_planilha2 = workbook2['Operacoes']

    max_linha = work_planilha2.max_row

    for i in range(2, max_linha + 1):
        date = datetime.strftime(work_planilha2.cell(row=i, column=5).value, "%m/%Y")
        work_planilha2.cell(row=i, column=5).value = date

    return workbook2.save('Planilha_Originaria_nova.xlsx')


# Saída
funcao_coluna_E()


def funcao_coluna_F():
    # Função que formata a data em (DD/MM/YYYY)
    workbook2 = load_workbook('Planilha_Originaria_nova.xlsx')
    work_planilha2 = workbook2['Operacoes']

    max_linha = work_planilha2.max_row

    for i in range(2, max_linha + 1):
        date = datetime.strftime(work_planilha2.cell(row=i, column=6).value, "%d/%m/%Y")
        work_planilha2.cell(row=i, column=6).value = date

    return workbook2.save('Planilha_Originaria_nova.xlsx')


# Saída
funcao_coluna_F()


def funcao_coluna_T():
    # Função que formata a data em (DD/MM/YYYY)
    workbook2 = load_workbook('Planilha_Originaria_nova.xlsx')
    work_planilha2 = workbook2['Operacoes']

    max_linha = work_planilha2.max_row

    for i in range(2, max_linha + 1):
        date = datetime.strftime(work_planilha2.cell(row=i, column=20).value, "%d/%m/%Y")
        work_planilha2.cell(row=i, column=20).value = date

    return workbook2.save('Planilha_Originaria_nova.xlsx')


# Saída
funcao_coluna_T()


def funcao_coluna_X():
    # Função que formata a data em (DD/MM/YYYY)
    workbook2 = load_workbook('Planilha_Originaria_nova.xlsx')
    work_planilha2 = workbook2['Operacoes']

    max_linha = work_planilha2.max_row

    for i in range(2, max_linha + 1):
        cell = work_planilha2.cell(row=i, column=24).value
        if cell is not None:
            date = datetime.strftime(work_planilha2.cell(row=i, column=24).value, "%d/%m/%Y")
            work_planilha2.cell(row=i, column=24).value = date

    return workbook2.save('Planilha_Originaria_nova.xlsx')


# Saída
funcao_coluna_X()


def funcao_coluna_O():
    # Função que formata os dados com 4 dígitos
    workbook2 = load_workbook('Planilha_Originaria_nova.xlsx')
    work_planilha2 = workbook2['Operacoes']

    max_linha = work_planilha2.max_row

    for i in range(2, max_linha + 1):
        cell_numero = float(work_planilha2.cell(row=i, column=15).value)
        cell_float = f'{cell_numero:.4f}'
        work_planilha2.cell(row=i, column=15).value = cell_float

    return workbook2.save('Planilha_Originaria_nova.xlsx')


# Saída
funcao_coluna_O()


def funcao_coluna_P():
    # Função que insere o R$
    workbook2 = load_workbook('Planilha_Originaria_nova.xlsx')
    work_planilha2 = workbook2['Operacoes']

    max_linha = work_planilha2.max_row

    for i in range(2, max_linha + 1):
        cell_numero = float(work_planilha2.cell(row=i, column=16).value)
        cell_float_2d = f'{cell_numero:.2f}'
        cell = ['R$', str(cell_float_2d)]
        work_planilha2.cell(row=i, column=16).value = ' '.join(cell)

    return workbook2.save('Planilha_Originaria_nova.xlsx')


# Saída
funcao_coluna_P()


def funcao_coluna_Q():
    # Função que insere o R$
    workbook2 = load_workbook('Planilha_Originaria_nova.xlsx')
    work_planilha2 = workbook2['Operacoes']

    max_linha = work_planilha2.max_row

    for i in range(2, max_linha + 1):
        cell_numero = float(work_planilha2.cell(row=i, column=17).value)
        cell_float_2d = f'{cell_numero:.2f}'
        cell = ['R$', str(cell_float_2d)]
        work_planilha2.cell(row=i, column=17).value = ' '.join(cell)

    return workbook2.save('Planilha_Originaria_nova.xlsx')


# Saída
funcao_coluna_Q()


def funcao_coluna_R():
    # Função que insere o R$
    workbook2 = load_workbook('Planilha_Originaria_nova.xlsx')
    work_planilha2 = workbook2['Operacoes']

    max_linha = work_planilha2.max_row

    for i in range(2, max_linha + 1):
        cell_numero = float(work_planilha2.cell(row=i, column=18).value)
        cell_float_2d = f'{cell_numero:.2f}'
        cell = ['R$', str(cell_float_2d)]
        work_planilha2.cell(row=i, column=18).value = ' '.join(cell)

    return workbook2.save('Planilha_Originaria_nova.xlsx')


# Saída
funcao_coluna_R()


def funcao_coluna_S():
    # Função que insere o R$
    workbook2 = load_workbook('Planilha_Originaria_nova.xlsx')
    work_planilha2 = workbook2['Operacoes']

    max_linha = work_planilha2.max_row

    for i in range(2, max_linha + 1):
        cell_numero = float(work_planilha2.cell(row=i, column=19).value)
        cell_float_2d = f'{cell_numero:.2f}'
        cell = ['R$', str(cell_float_2d)]
        work_planilha2.cell(row=i, column=19).value = ' '.join(cell)

    return workbook2.save('Planilha_Originaria_nova.xlsx')


# Saída
funcao_coluna_S()


def funcao_coluna_V():
    # Função que insere o R$
    workbook2 = load_workbook('Planilha_Originaria_nova.xlsx')
    work_planilha2 = workbook2['Operacoes']

    max_linha = work_planilha2.max_row

    for i in range(2, max_linha + 1):
        cell = work_planilha2.cell(row=i, column=22).value
        if cell is not None:
            cell_numero = float(work_planilha2.cell(row=i, column=22).value)
            cell_float_2d = f'{cell_numero:.2f}'
            cell = ['R$', str(cell_float_2d)]
            work_planilha2.cell(row=i, column=22).value = ' '.join(cell)

    return workbook2.save('Planilha_Originaria_nova.xlsx')


# Saída
funcao_coluna_V()
